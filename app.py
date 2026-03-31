import re
import io
import json
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import re
import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Analyse Comptable",
    page_icon="📊",
    layout="wide",
)

# ══════════════════════════════════════════════════════════════════════════════
# MENU PRINCIPAL — 5 boutons (Accueil + 4 modules)
# ══════════════════════════════════════════════════════════════════════════════
if "menu" not in st.session_state:
    st.session_state.menu = "🏠 Accueil"

st.markdown("## 📊 Analyse Comptable post-migration")

m1, m2, m3, m4, m5 = st.columns(5)
MENUS = [
    "🏠 Accueil",
    "📒 Grand Livre",
    "⚖️ Balance Auxiliaire",
    "📈 Balance Générale",
    "📗 Grand Livre Détaillé",
]
LABELS = [
    "🏠  Accueil",
    "📒  Grand Livre",
    "⚖️  Balance Auxiliaire",
    "📈  Balance Générale",
    "📗  Grand Livre Détaillé",
]
for col, key, label in zip([m1, m2, m3, m4, m5], MENUS, LABELS):
    with col:
        if st.button(label, use_container_width=True,
                     type="primary" if st.session_state.menu == key else "secondary"):
            st.session_state.menu = key
            st.rerun()

st.markdown("---")
menu = st.session_state.menu


# ══════════════════════════════════════════════════════════════════════════════
# UTILITAIRES COMMUNS
# ══════════════════════════════════════════════════════════════════════════════
def _make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_sheet(ws, df, title, tab_color):
    """Style professionnel uniforme pour un onglet Excel."""
    brd = _make_border()
    ws.title = title[:31]
    ws.sheet_properties.tabColor = tab_color
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    for ci, h in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font = hfill, hfont
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="EBF2FA" if ri % 2 == 0 else "FFFFFF")
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = brd
            c.fill = fill
            if isinstance(val, float):
                c.number_format = "#,##0.000"
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left")
    for ci, col in enumerate(df.columns, 1):
        w = max(len(str(col)), df[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 40)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def _to_float(s: str) -> float:
    """Convertit une chaîne numérique en float.
    Gère : espaces, séparateur milliers virgule (4,400.000) et format français (4 400,000).
    """
    s = str(s).strip().replace(" ", "")
    if not s or s == "-":
        return 0.0
    # virgule = séparateur de milliers, point = décimal  →  4,400.000
    if "." in s and "," in s:
        s = s.replace(",", "")
    # virgule seule = décimal (format français)  →  4400,000
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _highlight_ecarts(df):
    """Colorie en rouge/vert les colonnes d'écart dans un DataFrame styler."""
    styles = pd.DataFrame("", index=df.index, columns=df.columns)
    for col in [c for c in df.columns if c.startswith("Ecart_")]:
        styles[col] = df[col].apply(
            lambda v: "background-color: #FFCCCC" if isinstance(v, float) and v < -0.001
            else ("background-color: #CCFFCC" if isinstance(v, float) and v > 0.001 else "")
        )
    return styles


def _excel_color_ecarts(ws, comp, red_f, green_f):
    """Applique le coloriage rouge/vert sur les colonnes Ecart_ d'un onglet Excel."""
    ecart_idxs = [i + 1 for i, c in enumerate(comp.columns) if c.startswith("Ecart_")]
    for ri in range(2, len(comp) + 2):
        for ci in ecart_idxs:
            c = ws.cell(row=ri, column=ci)
            if c.value and abs(c.value) > 0.001:
                c.fill = red_f if c.value < 0 else green_f


def _excel_missing_sheet(wb, missing, cols_data, title, tab_color, la, lb, group_col, group_label_col):
    """Crée un onglet 'manquants' générique avec regroupement par entité."""
    brd = _make_border()
    ws = wb.create_sheet()
    ws.title = title[:31]
    ws.sheet_properties.tabColor = tab_color
    hfill    = PatternFill("solid", fgColor="1F4E79")
    hfont    = Font(color="FFFFFF", bold=True, size=11)
    red_fill = PatternFill("solid", fgColor="FFE0E0")
    blu_fill = PatternFill("solid", fgColor="E0EEFF")

    ws.cell(row=1, column=1, value="Entité / Nom").fill = hfill
    ws.cell(row=1, column=1).font = hfont
    ws.cell(row=1, column=1).border = brd
    for ci, h in enumerate(cols_data, 2):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font = hfill, hfont
        c.alignment = Alignment(horizontal="center")
        c.border = brd
    ws.row_dimensions[1].height = 20

    sup_fill = PatternFill("solid", fgColor="D9E1F2")
    sup_font = Font(bold=True, size=11)
    cur = 2

    # Index de la colonne "Absent dans" dans cols_data
    absent_idx = cols_data.index("Absent dans") + 2 if "Absent dans" in cols_data else None

    for grp_code, grp in missing.groupby(group_col, sort=True):
        # Ligne titre groupe
        n_cols = len(cols_data) + 1
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=n_cols)
        lbl = grp[group_label_col].iloc[0] if group_label_col in grp.columns else ""
        sc = ws.cell(row=cur, column=1, value=f"  {grp_code}  –  {lbl}")
        sc.fill, sc.font = sup_fill, sup_font
        sc.alignment = Alignment(horizontal="left", vertical="center")
        sc.border = brd
        ws.row_dimensions[cur].height = 18
        cur += 1
        for _, drow in grp.iterrows():
            # Détermine la couleur selon "Absent dans"
            if absent_idx:
                rf = red_fill if drow.get("Absent dans", "") == lb else blu_fill
            else:
                rf = red_fill
            ws.cell(row=cur, column=1, value="").fill = rf
            for ci, col in enumerate(cols_data, 2):
                val = drow[col] if col in drow.index else ""
                c = ws.cell(row=cur, column=ci, value=val)
                c.fill, c.border = rf, brd
                if isinstance(val, float):
                    c.number_format = "#,##0.000"
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.alignment = Alignment(horizontal="left")
            cur += 1
        cur += 1

    ws.column_dimensions["A"].width = 35
    for ci, col in enumerate(cols_data, 2):
        w = max(len(col), missing[col].astype(str).str.len().max() if col in missing.columns else 10)
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 40)
    ws.freeze_panes = "A2"

    lr = cur + 1
    ws.cell(row=lr,   column=1, value="Légende :").font = Font(bold=True)
    ws.cell(row=lr+1, column=1, value=f"  Absent dans {lb}").fill = red_fill
    ws.cell(row=lr+2, column=1, value=f"  Absent dans {la}").fill = blu_fill
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# ACCUEIL
# ══════════════════════════════════════════════════════════════════════════════
if menu == "🏠 Accueil":
    st.title("📊 Analyse Comptable post-migration")
    st.markdown("Sélectionnez un module dans le menu ci-dessus.")
    st.markdown("---")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
### 📒 Grand Livre
Comparaison de deux fichiers **Grand Livre fournisseurs** au format TXT
(colonnes fixes, une ligne par transaction, comptes `F\\d+`).

**Fonctionnalités :**
- Comparaison agrégée Débit / Crédit par fournisseur
- Détection des documents manquants
- Export Excel multi-onglets
        """)
        if st.button("Ouvrir Grand Livre →", use_container_width=True):
            st.session_state.menu = "📒 Grand Livre"
            st.rerun()

        st.markdown("---")
        st.markdown("""
### 📈 Balance Générale
Comparaison de deux fichiers **Balance Générale** au format TXT
(tableau `|Compte|` / `|Description|`, 3 sections : Balance antérieure, Mouvement, Solde).

**Fonctionnalités :**
- Comparaison agrégée Débit / Crédit par compte
- Détection des comptes manquants
- Écarts colorisés (rouge / vert)
- Export Excel multi-onglets
        """)
        if st.button("Ouvrir Balance Générale →", use_container_width=True):
            st.session_state.menu = "📈 Balance Générale"
            st.rerun()

    with c2:
        st.markdown("""
### ⚖️ Balance Auxiliaire
Comparaison de deux fichiers **Balance Auxiliaire** au format TXT
(2 lignes par fournisseur séparées par `|`).

**Fonctionnalités :**
- Balance antérieure, mouvements, balance finale, solde
- Fournisseurs communs, manquants, écarts
- Labels personnalisables (Fichier A / Fichier B)
- Export Excel multi-onglets
        """)
        if st.button("Ouvrir Balance Auxiliaire →", use_container_width=True):
            st.session_state.menu = "⚖️ Balance Auxiliaire"
            st.rerun()

        st.markdown("---")
        st.markdown("""
### 📗 Grand Livre Détaillé
Comparaison de deux fichiers **Grand Livre** au format TXT
(transactions ligne à ligne séparées par `|`, regroupées par compte).

**Fonctionnalités :**
- Comparaison agrégée Débit / Crédit / Solde final par compte
- Détection des références manquantes par compte
- Filtre par compte sur les détails
- Export Excel multi-onglets
        """)
        if st.button("Ouvrir Grand Livre Détaillé →", use_container_width=True):
            st.session_state.menu = "📗 Grand Livre Détaillé"
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 1 — GRAND LIVRE (fournisseurs F\d+)
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📒 Grand Livre":

    st.title("📒 Comparaison Grand Livre")

    @st.cache_data
    def parse_grand_livre(file_bytes: bytes, label: str = "fichier") -> pd.DataFrame:
        rows = []
        current_code = current_name = ""
        for line in file_bytes.decode("utf-8", errors="ignore").splitlines():
            line = line.replace("\r", "")
            m = re.match(r"^(F\d+)\s+(.*)", line)
            if m:
                current_code = m.group(1).strip()
                current_name = m.group(2).strip()
                continue
            if not re.match(r"^\d{2}/\d{2}/\d{2}", line):
                continue
            if re.search(r"(Tot du|Cumuls au|cumuls au)", line):
                continue
            if len(line) < 50:
                continue
            solde  = line[-14:].strip()
            credit = line[-28:-14].strip()
            debit  = line[-42:-28].strip()
            rest   = line[:-42].strip()
            parts  = rest.split()
            if len(parts) < 4:
                continue
            rows.append({
                "Fournisseur": current_code,
                "Nom":         current_name,
                "Date":        parts[0],
                "Document":    parts[1],
                "Type":        parts[2],
                "Reference":   parts[3],
                "Debit":       debit,
                "Credit":      credit,
                "Solde":       solde,
            })
        if not rows:
            st.warning(f"⚠️ **{label}** : aucune transaction détectée.")
            return pd.DataFrame(columns=["Fournisseur", "Nom", "Date", "Document",
                                          "Type", "Reference", "Debit", "Credit", "Solde"])
        df = pd.DataFrame(rows)
        for col in ["Debit", "Credit", "Solde"]:
            df[col] = (df[col].astype(str)
                       .str.replace(r"\s+", "", regex=True)
                       .str.replace(",", ".", regex=False))
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df

    def gl1_compute_missing(df1, df2, la, lb):
        all_suppliers = (pd.concat([df1[["Fournisseur", "Nom"]], df2[["Fournisseur", "Nom"]]])
                         .drop_duplicates("Fournisseur").sort_values("Fournisseur"))
        records = []
        for _, sup_row in all_suppliers.iterrows():
            code = sup_row["Fournisseur"]
            nom  = sup_row["Nom"]
            docs1 = set(df1.loc[df1["Fournisseur"] == code, "Document"])
            docs2 = set(df2.loc[df2["Fournisseur"] == code, "Document"])
            for doc in sorted(docs1 - docs2):
                r = df1[(df1["Fournisseur"] == code) & (df1["Document"] == doc)].iloc[0]
                records.append({"Fournisseur": code, "Nom": nom, "Document": doc,
                                "Présent dans": la, "Absent dans": lb,
                                "Date": r["Date"], "Type": r["Type"],
                                "Reference": r["Reference"],
                                "Debit": r["Debit"], "Credit": r["Credit"]})
            for doc in sorted(docs2 - docs1):
                r = df2[(df2["Fournisseur"] == code) & (df2["Document"] == doc)].iloc[0]
                records.append({"Fournisseur": code, "Nom": nom, "Document": doc,
                                "Présent dans": lb, "Absent dans": la,
                                "Date": r["Date"], "Type": r["Type"],
                                "Reference": r["Reference"],
                                "Debit": r["Debit"], "Credit": r["Credit"]})
        if not records:
            return pd.DataFrame()
        return pd.DataFrame(records).sort_values(["Fournisseur", "Présent dans", "Document"])

    def gl1_build_excel(df1, df2, comp, missing, la, lb):
        wb  = Workbook()
        red_f   = PatternFill("solid", fgColor="FFCCCC")
        green_f = PatternFill("solid", fgColor="CCFFCC")

        ws1 = wb.active
        _style_sheet(ws1, comp, "Comparaison", "1F4E79")
        _excel_color_ecarts(ws1, comp, red_f, green_f)

        if not missing.empty:
            _excel_missing_sheet(wb, missing,
                                  ["Document", "Présent dans", "Absent dans",
                                   "Date", "Type", "Reference", "Debit", "Credit"],
                                  "Documents manquants", "C00000", la, lb,
                                  "Fournisseur", "Nom")

        ws3 = wb.create_sheet()
        _style_sheet(ws3, df1, f"Détail {la}", "2E75B6")
        ws4 = wb.create_sheet()
        _style_sheet(ws4, df2, f"Détail {lb}", "70AD47")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Upload ──────────────────────────────────────────────────────────────
    c1, c2 = st.columns(2)
    with c1:
        f1 = st.file_uploader("📂 Fichier A", type=["txt"], key="gl1_f1")
        LA = st.text_input("Nom du fichier A", value="Fichier A", key="gl1_la")
    with c2:
        f2 = st.file_uploader("📂 Fichier B", type=["txt"], key="gl1_f2")
        LB = st.text_input("Nom du fichier B", value="Fichier B", key="gl1_lb")

    if f1 and f2:
        with st.spinner("Parsing en cours..."):
            df1 = parse_grand_livre(f1.read(), LA)
            df2 = parse_grand_livre(f2.read(), LB)

        if df1.empty or df2.empty:
            st.error("Impossible de parser un ou plusieurs fichiers.")
            st.stop()

        sa, sb = f"_{LA}", f"_{LB}"
        agg1 = df1.groupby(["Fournisseur", "Nom"])[["Debit", "Credit"]].sum().reset_index()
        agg2 = df2.groupby(["Fournisseur", "Nom"])[["Debit", "Credit"]].sum().reset_index()
        comp = pd.merge(agg1, agg2, on="Fournisseur", how="outer", suffixes=(sa, sb))
        comp["Nom"] = comp[f"Nom{sa}"].fillna(comp[f"Nom{sb}"])
        comp = comp.drop(columns=[f"Nom{sa}", f"Nom{sb}"])
        comp = comp[["Fournisseur", "Nom",
                      f"Debit{sa}", f"Credit{sa}",
                      f"Debit{sb}", f"Credit{sb}"]].fillna(0)
        comp["Ecart_Debit"]  = comp[f"Debit{sb}"]  - comp[f"Debit{sa}"]
        comp["Ecart_Credit"] = comp[f"Credit{sb}"] - comp[f"Credit{sa}"]

        missing = gl1_compute_missing(df1, df2, LA, LB)

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric(f"Fournisseurs {LA}", df1["Fournisseur"].nunique())
        k2.metric(f"Fournisseurs {LB}", df2["Fournisseur"].nunique())
        k3.metric(f"Lignes {LA}", len(df1))
        k4.metric(f"Lignes {LB}", len(df2))
        k5.metric("Documents manquants", len(missing) if not missing.empty else 0,
                  delta="⚠️" if not missing.empty else None, delta_color="inverse")

        tab1, tab2, tab3, tab4 = st.tabs([
            "📊 Comparaison agrégée",
            "🔍 Documents manquants",
            f"📄 Détail {LA}",
            f"📄 Détail {LB}",
        ])
        fmt = {c: "{:,.3f}" for c in comp.columns if comp[c].dtype == float}

        with tab1:
            st.dataframe(comp.style.format(fmt).apply(_highlight_ecarts, axis=None),
                         use_container_width=True)

        with tab2:
            if missing.empty:
                st.success("✅ Aucun document manquant.")
            else:
                st.info(f"**{len(missing)} document(s) manquant(s)** sur "
                        f"**{missing['Fournisseur'].nunique()} fournisseur(s)**")
                for sup_code, grp in missing.groupby("Fournisseur", sort=True):
                    nom    = grp["Nom"].iloc[0]
                    only_a = grp[grp["Absent dans"] == LB]
                    only_b = grp[grp["Absent dans"] == LA]
                    label  = (f"**{sup_code}** – {nom}  "
                              + (f"🔴 {len(only_a)} absent(s) dans {LB}  " if not only_a.empty else "")
                              + (f"🔵 {len(only_b)} absent(s) dans {LA}"   if not only_b.empty else ""))
                    with st.expander(label, expanded=False):
                        cols_d = ["Document", "Date", "Type", "Reference", "Debit", "Credit"]
                        if not only_a.empty:
                            st.markdown(f"🔴 **Présents dans {LA} — Absents dans {LB}**")
                            st.dataframe(only_a[cols_d].reset_index(drop=True),
                                         use_container_width=True)
                        if not only_b.empty:
                            st.markdown(f"🔵 **Présents dans {LB} — Absents dans {LA}**")
                            st.dataframe(only_b[cols_d].reset_index(drop=True),
                                         use_container_width=True)

        with tab3:
            st.dataframe(df1.style.format(
                {c: "{:,.3f}" for c in df1.columns if df1[c].dtype == float}
            ), use_container_width=True)

        with tab4:
            st.dataframe(df2.style.format(
                {c: "{:,.3f}" for c in df2.columns if df2[c].dtype == float}
            ), use_container_width=True)

        st.divider()
        st.download_button(
            label="📥 Télécharger Excel",
            data=gl1_build_excel(df1, df2, comp, missing, LA, LB),
            file_name="comparaison_grand_livre.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 2 — BALANCE AUXILIAIRE
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "⚖️ Balance Auxiliaire":

    st.title("⚖️ Comparaison Balance Auxiliaire")

    @st.cache_data
    def parse_balance(file_bytes: bytes, label: str = "fichier") -> pd.DataFrame:
        def extract_trailing_number(seg: str) -> float:
            m = re.search(r"([\d.]+)\s*$", seg.strip())
            return _to_float(m.group(1)) if m else 0.0

        def clean_name(raw: str) -> str:
            return re.sub(r"\s+[\d,.]+\s*$", "", raw).strip()

        rows  = []
        lines = file_bytes.decode("utf-8", errors="ignore").splitlines()
        lines = [l.replace("\r", "") for l in lines]
        i = 0
        while i < len(lines):
            line = lines[i]
            if re.match(r"^[Ff]\d+", line):
                fline = line
                j = i + 1
                while j < len(lines) and not lines[j].strip():
                    j += 1
                nline  = lines[j] if j < len(lines) else ""
                fparts = fline.split("|")
                nparts = nline.split("|")
                code_m = re.match(r"^([Ff]\d+)", fparts[0])
                code   = code_m.group(1).upper() if code_m else ""
                name   = clean_name(nparts[0]) if nparts else ""
                rows.append({
                    "Fournisseur":   code,
                    "Nom":           name,
                    "BalAnt_Debit":  extract_trailing_number(fparts[0]),
                    "BalAnt_Credit": extract_trailing_number(nparts[0]) if nparts else 0.0,
                    "Mvt_Debit":     _to_float(fparts[1]) if len(fparts) > 1 else 0.0,
                    "Mvt_Credit":    _to_float(nparts[1]) if len(nparts) > 1 else 0.0,
                    "Bal_Debit":     _to_float(fparts[2]) if len(fparts) > 2 else 0.0,
                    "Bal_Credit":    _to_float(nparts[2]) if len(nparts) > 2 else 0.0,
                    "Solde_Debit":   _to_float(fparts[3]) if len(fparts) > 3 else 0.0,
                    "Solde_Credit":  _to_float(nparts[3]) if len(nparts) > 3 else 0.0,
                })
                i = j + 1
            else:
                i += 1

        if not rows:
            st.warning(f"⚠️ **{label}** : aucun fournisseur détecté.")
            return pd.DataFrame(columns=["Fournisseur", "Nom",
                                          "BalAnt_Debit", "BalAnt_Credit",
                                          "Mvt_Debit", "Mvt_Credit",
                                          "Bal_Debit", "Bal_Credit",
                                          "Solde_Debit", "Solde_Credit"])
        return pd.DataFrame(rows)

    def ba_compute_common(df1, df2, la, lb):
        codes  = set(df1["Fournisseur"]) & set(df2["Fournisseur"])
        sa, sb = f"_{la}", f"_{lb}"
        merged = pd.merge(df1[df1["Fournisseur"].isin(codes)],
                          df2[df2["Fournisseur"].isin(codes)],
                          on="Fournisseur", suffixes=(sa, sb))
        merged["Nom"] = merged[f"Nom{sa}"].fillna(merged[f"Nom{sb}"])
        merged = merged.drop(columns=[f"Nom{sa}", f"Nom{sb}"])
        cols = ["Fournisseur", "Nom"] + [c for c in merged.columns
                                          if c not in ("Fournisseur", "Nom")]
        return merged[cols].sort_values("Fournisseur").reset_index(drop=True)

    def ba_compute_missing(df1, df2, la, lb):
        codes1, codes2 = set(df1["Fournisseur"]), set(df2["Fournisseur"])
        records = []
        for code in sorted(codes1 - codes2):
            r = df1[df1["Fournisseur"] == code].iloc[0]
            records.append({"Fournisseur": code, "Nom": r["Nom"],
                            "Présent dans": la, "Absent dans": lb,
                            "BalAnt_Debit": r["BalAnt_Debit"], "BalAnt_Credit": r["BalAnt_Credit"],
                            "Mvt_Debit": r["Mvt_Debit"], "Mvt_Credit": r["Mvt_Credit"],
                            "Bal_Debit": r["Bal_Debit"], "Bal_Credit": r["Bal_Credit"],
                            "Solde_Debit": r["Solde_Debit"], "Solde_Credit": r["Solde_Credit"]})
        for code in sorted(codes2 - codes1):
            r = df2[df2["Fournisseur"] == code].iloc[0]
            records.append({"Fournisseur": code, "Nom": r["Nom"],
                            "Présent dans": lb, "Absent dans": la,
                            "BalAnt_Debit": r["BalAnt_Debit"], "BalAnt_Credit": r["BalAnt_Credit"],
                            "Mvt_Debit": r["Mvt_Debit"], "Mvt_Credit": r["Mvt_Credit"],
                            "Bal_Debit": r["Bal_Debit"], "Bal_Credit": r["Bal_Credit"],
                            "Solde_Debit": r["Solde_Debit"], "Solde_Credit": r["Solde_Credit"]})
        if not records:
            return pd.DataFrame()
        return pd.DataFrame(records).sort_values(["Absent dans", "Fournisseur"])

    def ba_build_excel(df1, df2, comp, missing, common, la, lb):
        wb      = Workbook()
        red_f   = PatternFill("solid", fgColor="FFCCCC")
        green_f = PatternFill("solid", fgColor="CCFFCC")

        ws1 = wb.active
        _style_sheet(ws1, comp, "Comparaison", "1F4E79")
        _excel_color_ecarts(ws1, comp, red_f, green_f)

        ws_com = wb.create_sheet()
        _style_sheet(ws_com, common, "Fournisseurs communs", "375623")

        if not missing.empty:
            _excel_missing_sheet(wb, missing,
                                  ["Présent dans", "Absent dans",
                                   "BalAnt_Debit", "BalAnt_Credit",
                                   "Mvt_Debit", "Mvt_Credit",
                                   "Bal_Debit", "Bal_Credit",
                                   "Solde_Debit", "Solde_Credit"],
                                  "Fournisseurs manquants", "C00000", la, lb,
                                  "Fournisseur", "Nom")

        ws4 = wb.create_sheet()
        _style_sheet(ws4, df1, f"Données {la}", "2E75B6")
        ws5 = wb.create_sheet()
        _style_sheet(ws5, df2, f"Données {lb}", "70AD47")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Upload ──────────────────────────────────────────────────────────────
    c1, c2 = st.columns(2)
    with c1:
        f1 = st.file_uploader("📂 Fichier A", type=["txt"], key="ba_f1")
        LA = st.text_input("Nom du fichier A", value="Fichier A", key="ba_la")
    with c2:
        f2 = st.file_uploader("📂 Fichier B", type=["txt"], key="ba_f2")
        LB = st.text_input("Nom du fichier B", value="Fichier B", key="ba_lb")

    if f1 and f2:
        with st.spinner("Parsing en cours..."):
            df1 = parse_balance(f1.read(), LA)
            df2 = parse_balance(f2.read(), LB)

        if df1.empty or df2.empty:
            st.error("Impossible de parser un ou plusieurs fichiers.")
            st.stop()

        sa, sb = f"_{LA}", f"_{LB}"
        comp = pd.merge(df1, df2, on="Fournisseur", how="outer", suffixes=(sa, sb))
        comp["Nom"] = comp[f"Nom{sa}"].fillna(comp[f"Nom{sb}"])
        comp = comp.drop(columns=[f"Nom{sa}", f"Nom{sb}"]).fillna(0)
        base_cols = ["Fournisseur", "Nom",
                     f"BalAnt_Debit{sa}", f"BalAnt_Credit{sa}",
                     f"BalAnt_Debit{sb}", f"BalAnt_Credit{sb}",
                     f"Mvt_Debit{sa}",    f"Mvt_Credit{sa}",
                     f"Mvt_Debit{sb}",    f"Mvt_Credit{sb}",
                     f"Bal_Debit{sa}",    f"Bal_Credit{sa}",
                     f"Bal_Debit{sb}",    f"Bal_Credit{sb}",
                     f"Solde_Debit{sa}",  f"Solde_Credit{sa}",
                     f"Solde_Debit{sb}",  f"Solde_Credit{sb}"]
        comp = comp[base_cols]
        comp["Ecart_BalAnt_Debit"]  = comp[f"BalAnt_Debit{sb}"]  - comp[f"BalAnt_Debit{sa}"]
        comp["Ecart_BalAnt_Credit"] = comp[f"BalAnt_Credit{sb}"] - comp[f"BalAnt_Credit{sa}"]
        comp["Ecart_Mvt_Debit"]     = comp[f"Mvt_Debit{sb}"]     - comp[f"Mvt_Debit{sa}"]
        comp["Ecart_Mvt_Credit"]    = comp[f"Mvt_Credit{sb}"]    - comp[f"Mvt_Credit{sa}"]
        comp["Ecart_Bal_Debit"]     = comp[f"Bal_Debit{sb}"]     - comp[f"Bal_Debit{sa}"]
        comp["Ecart_Bal_Credit"]    = comp[f"Bal_Credit{sb}"]    - comp[f"Bal_Credit{sa}"]
        comp["Ecart_Solde_Debit"]   = comp[f"Solde_Debit{sb}"]   - comp[f"Solde_Debit{sa}"]
        comp["Ecart_Solde_Credit"]  = comp[f"Solde_Credit{sb}"]  - comp[f"Solde_Credit{sa}"]

        missing = ba_compute_missing(df1, df2, LA, LB)
        common  = ba_compute_common(df1, df2, LA, LB)

        nb_ecart = len(comp[
            (comp["Ecart_BalAnt_Debit"].abs()  > 0.01) |
            (comp["Ecart_BalAnt_Credit"].abs() > 0.01) |
            (comp["Ecart_Mvt_Debit"].abs()     > 0.01) |
            (comp["Ecart_Mvt_Credit"].abs()    > 0.01)
        ])

        k1, k2, k3, k4, k5, k6 = st.columns(6)
        k1.metric(f"Fournisseurs {LA}", len(df1))
        k2.metric(f"Fournisseurs {LB}", len(df2))
        k3.metric("Communs", len(set(df1["Fournisseur"]) & set(df2["Fournisseur"])))
        k4.metric(f"Uniq. {LA}", len(set(df1["Fournisseur"]) - set(df2["Fournisseur"])))
        k5.metric(f"Uniq. {LB}", len(set(df2["Fournisseur"]) - set(df1["Fournisseur"])))
        k6.metric("Avec écart", nb_ecart,
                  delta="⚠️" if nb_ecart > 0 else None, delta_color="inverse")

        fmt = {c: "{:,.3f}" for c in comp.columns if comp[c].dtype == float}

        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📊 Comparaison",
            "🤝 Fournisseurs communs",
            "🔍 Fournisseurs manquants",
            f"📄 Données {LA}",
            f"📄 Données {LB}",
        ])

        with tab1:
            fc1, fc2 = st.columns([2, 1])
            with fc2:
                only_ecart = st.toggle("Afficher uniquement les écarts", value=False,
                                        key="ba_toggle")
            display = comp.copy()
            if only_ecart:
                display = display[
                    (display["Ecart_BalAnt_Debit"].abs()  > 0.01) |
                    (display["Ecart_BalAnt_Credit"].abs() > 0.01) |
                    (display["Ecart_Mvt_Debit"].abs()     > 0.01) |
                    (display["Ecart_Mvt_Credit"].abs()    > 0.01)
                ]
                st.caption(f"{len(display)} fournisseur(s) avec écart")
            st.dataframe(display.style.format(fmt).apply(_highlight_ecarts, axis=None),
                         use_container_width=True)

        with tab2:
            st.caption(f"{len(common)} fournisseurs présents dans les deux fichiers")
            st.dataframe(common.style.format(
                {c: "{:,.3f}" for c in common.columns if common[c].dtype == float}
            ), use_container_width=True)

        with tab3:
            if missing.empty:
                st.success("✅ Tous les fournisseurs sont présents dans les deux fichiers.")
            else:
                only_a = missing[missing["Absent dans"] == LB]
                only_b = missing[missing["Absent dans"] == LA]
                st.info(
                    f"**{len(missing)} fournisseur(s) manquant(s)** — "
                    f"🔴 {len(only_a)} absent(s) dans {LB} · "
                    f"🔵 {len(only_b)} absent(s) dans {LA}"
                )
                fmt_m = {c: "{:,.3f}" for c in missing.columns if missing[c].dtype == float}
                if not only_a.empty:
                    st.markdown(f"#### 🔴 Présents dans {LA} — Absents dans {LB}")
                    st.dataframe(only_a.drop(columns=["Présent dans", "Absent dans"])
                                 .reset_index(drop=True).style.format(fmt_m),
                                 use_container_width=True)
                if not only_b.empty:
                    st.markdown(f"#### 🔵 Présents dans {LB} — Absents dans {LA}")
                    st.dataframe(only_b.drop(columns=["Présent dans", "Absent dans"])
                                 .reset_index(drop=True).style.format(fmt_m),
                                 use_container_width=True)

        with tab4:
            st.dataframe(df1.style.format(
                {c: "{:,.3f}" for c in df1.columns if df1[c].dtype == float}
            ), use_container_width=True)

        with tab5:
            st.dataframe(df2.style.format(
                {c: "{:,.3f}" for c in df2.columns if df2[c].dtype == float}
            ), use_container_width=True)

        st.divider()
        st.download_button(
            label="📥 Télécharger Excel (5 onglets)",
            data=ba_build_excel(df1, df2, comp, missing, common, LA, LB),
            file_name="comparaison_balance_auxiliaire.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 3 — BALANCE GÉNÉRALE
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📈 Balance Générale":

    st.title("📈 Comparaison Balance Générale")

    @st.cache_data
    def parse_balance_generale(file_bytes: bytes, label: str = "fichier") -> pd.DataFrame:
        """
        Format : 2 lignes par compte.
          |10100000   |  0.000|       |  0.000|       |       |       |   ← débits
          |CAPITAL SOCIAL |       |  0.000|       |  1,939,390.000|...| ← crédits
        """
        lines = file_bytes.decode("utf-8", errors="ignore").splitlines()
        lines = [l.replace("\r", "") for l in lines]

        def parse_pipe_values(line: str):
            cells = line.split("|")
            vals  = []
            for cell in cells[2:]:
                cell = cell.strip()
                vals.append(_to_float(cell) if cell else None)
            return vals

        rows = []
        i = 0
        while i < len(lines):
            line = lines[i]
            m = re.match(r"^\|(\d{8})\s*\|", line)
            if m:
                code  = m.group(1)
                vals1 = parse_pipe_values(line)

                j = i + 1
                while j < len(lines) and not lines[j].strip():
                    j += 1
                if j < len(lines) and re.match(r"^\|[A-Za-zÀ-ÿ\s/\.\-\']", lines[j]):
                    dm = re.match(r"^\|([^|]+)\|", lines[j])
                    description = dm.group(1).strip() if dm else ""
                    vals2 = parse_pipe_values(lines[j])
                    i = j + 1
                else:
                    description = ""
                    vals2 = [None] * 6
                    i += 1

                def pick(lst, idx):
                    if lst and idx < len(lst) and lst[idx] is not None:
                        return lst[idx]
                    return 0.0

                rows.append({
                    "Compte":        code,
                    "Description":   description,
                    "BalAnt_Debit":  pick(vals1, 0),
                    "BalAnt_Credit": pick(vals2, 1),
                    "Mvt_Debit":     pick(vals1, 2),
                    "Mvt_Credit":    pick(vals2, 3),
                    "Solde_Debit":   pick(vals1, 4),
                    "Solde_Credit":  pick(vals2, 5),
                })
            else:
                i += 1

        if not rows:
            st.warning(f"⚠️ **{label}** : aucun compte détecté.")
            return pd.DataFrame(columns=["Compte", "Description",
                                          "BalAnt_Debit", "BalAnt_Credit",
                                          "Mvt_Debit", "Mvt_Credit",
                                          "Solde_Debit", "Solde_Credit"])
        return pd.DataFrame(rows).sort_values("Compte").reset_index(drop=True)

    def bg_compute_missing(df1, df2, la, lb):
        codes1, codes2 = set(df1["Compte"]), set(df2["Compte"])
        records = []
        for code in sorted(codes1 - codes2):
            r = df1[df1["Compte"] == code].iloc[0]
            records.append({"Compte": code, "Description": r["Description"],
                            "Présent dans": la, "Absent dans": lb,
                            "BalAnt_Debit": r["BalAnt_Debit"], "BalAnt_Credit": r["BalAnt_Credit"],
                            "Mvt_Debit": r["Mvt_Debit"], "Mvt_Credit": r["Mvt_Credit"],
                            "Solde_Debit": r["Solde_Debit"], "Solde_Credit": r["Solde_Credit"]})
        for code in sorted(codes2 - codes1):
            r = df2[df2["Compte"] == code].iloc[0]
            records.append({"Compte": code, "Description": r["Description"],
                            "Présent dans": lb, "Absent dans": la,
                            "BalAnt_Debit": r["BalAnt_Debit"], "BalAnt_Credit": r["BalAnt_Credit"],
                            "Mvt_Debit": r["Mvt_Debit"], "Mvt_Credit": r["Mvt_Credit"],
                            "Solde_Debit": r["Solde_Debit"], "Solde_Credit": r["Solde_Credit"]})
        if not records:
            return pd.DataFrame()
        return pd.DataFrame(records).sort_values(["Absent dans", "Compte"]).reset_index(drop=True)

    def bg_build_excel(df1, df2, comp, missing, la, lb):
        wb      = Workbook()
        red_f   = PatternFill("solid", fgColor="FFCCCC")
        green_f = PatternFill("solid", fgColor="CCFFCC")

        ws1 = wb.active
        _style_sheet(ws1, comp, "Comparaison", "1F4E79")
        _excel_color_ecarts(ws1, comp, red_f, green_f)

        codes_communs = set(df1["Compte"]) & set(df2["Compte"])
        df_commun = comp[comp["Compte"].isin(codes_communs)].copy()
        ws_com = wb.create_sheet()
        _style_sheet(ws_com, df_commun, "Comptes communs", "375623")
        _excel_color_ecarts(ws_com, df_commun.reset_index(drop=True), red_f, green_f)

        if not missing.empty:
            _excel_missing_sheet(wb, missing,
                                  ["Présent dans", "Absent dans",
                                   "BalAnt_Debit", "BalAnt_Credit",
                                   "Mvt_Debit", "Mvt_Credit",
                                   "Solde_Debit", "Solde_Credit"],
                                  "Comptes manquants", "C00000", la, lb,
                                  "Compte", "Description")

        ws4 = wb.create_sheet()
        _style_sheet(ws4, df1, f"Données {la}", "2E75B6")
        ws5 = wb.create_sheet()
        _style_sheet(ws5, df2, f"Données {lb}", "70AD47")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Upload ──────────────────────────────────────────────────────────────
    c1, c2 = st.columns(2)
    with c1:
        f1 = st.file_uploader("📂 Fichier A (Balance Générale)", type=["txt"], key="bg_f1")
        LA = st.text_input("Nom du fichier A", value="Fichier A", key="bg_la")
    with c2:
        f2 = st.file_uploader("📂 Fichier B (Balance Générale)", type=["txt"], key="bg_f2")
        LB = st.text_input("Nom du fichier B", value="Fichier B", key="bg_lb")

    if f1 and f2:
        with st.spinner("Parsing en cours..."):
            df1 = parse_balance_generale(f1.read(), LA)
            df2 = parse_balance_generale(f2.read(), LB)

        if df1.empty or df2.empty:
            st.error("Impossible de parser un ou plusieurs fichiers.")
            st.stop()

        sa, sb = f"_{LA}", f"_{LB}"
        comp = pd.merge(df1, df2, on="Compte", how="outer", suffixes=(sa, sb))
        comp["Description"] = comp[f"Description{sa}"].fillna(comp[f"Description{sb}"])
        comp = comp.drop(columns=[f"Description{sa}", f"Description{sb}"]).fillna(0)
        base_cols = ["Compte", "Description",
                     f"BalAnt_Debit{sa}", f"BalAnt_Credit{sa}",
                     f"BalAnt_Debit{sb}", f"BalAnt_Credit{sb}",
                     f"Mvt_Debit{sa}",    f"Mvt_Credit{sa}",
                     f"Mvt_Debit{sb}",    f"Mvt_Credit{sb}",
                     f"Solde_Debit{sa}",  f"Solde_Credit{sa}",
                     f"Solde_Debit{sb}",  f"Solde_Credit{sb}"]
        comp = comp[base_cols]
        comp["Ecart_BalAnt_Debit"]  = comp[f"BalAnt_Debit{sb}"]  - comp[f"BalAnt_Debit{sa}"]
        comp["Ecart_BalAnt_Credit"] = comp[f"BalAnt_Credit{sb}"] - comp[f"BalAnt_Credit{sa}"]
        comp["Ecart_Mvt_Debit"]     = comp[f"Mvt_Debit{sb}"]     - comp[f"Mvt_Debit{sa}"]
        comp["Ecart_Mvt_Credit"]    = comp[f"Mvt_Credit{sb}"]    - comp[f"Mvt_Credit{sa}"]
        comp["Ecart_Solde_Debit"]   = comp[f"Solde_Debit{sb}"]   - comp[f"Solde_Debit{sa}"]
        comp["Ecart_Solde_Credit"]  = comp[f"Solde_Credit{sb}"]  - comp[f"Solde_Credit{sa}"]
        comp = comp.sort_values("Compte").reset_index(drop=True)

        missing       = bg_compute_missing(df1, df2, LA, LB)
        codes_communs = set(df1["Compte"]) & set(df2["Compte"])

        nb_ecart = len(comp[
            (comp["Ecart_BalAnt_Debit"].abs()  > 0.001) |
            (comp["Ecart_BalAnt_Credit"].abs() > 0.001) |
            (comp["Ecart_Mvt_Debit"].abs()     > 0.001) |
            (comp["Ecart_Mvt_Credit"].abs()    > 0.001) |
            (comp["Ecart_Solde_Debit"].abs()   > 0.001) |
            (comp["Ecart_Solde_Credit"].abs()  > 0.001)
        ])

        k1, k2, k3, k4, k5, k6 = st.columns(6)
        k1.metric(f"Comptes {LA}", len(df1))
        k2.metric(f"Comptes {LB}", len(df2))
        k3.metric("Communs", len(codes_communs))
        k4.metric(f"Uniq. {LA}", len(set(df1["Compte"]) - set(df2["Compte"])))
        k5.metric(f"Uniq. {LB}", len(set(df2["Compte"]) - set(df1["Compte"])))
        k6.metric("Avec écart", nb_ecart,
                  delta="⚠️" if nb_ecart > 0 else None, delta_color="inverse")

        fmt = {c: "{:,.3f}" for c in comp.columns if comp[c].dtype == float}

        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📊 Comparaison",
            "🤝 Comptes communs",
            "🔍 Comptes manquants",
            f"📄 Données {LA}",
            f"📄 Données {LB}",
        ])

        with tab1:
            fc1, fc2 = st.columns([2, 1])
            with fc2:
                only_ecart = st.toggle("Afficher uniquement les écarts", value=False,
                                        key="bg_toggle")
            display = comp.copy()
            if only_ecart:
                display = display[
                    (display["Ecart_BalAnt_Debit"].abs()  > 0.001) |
                    (display["Ecart_BalAnt_Credit"].abs() > 0.001) |
                    (display["Ecart_Mvt_Debit"].abs()     > 0.001) |
                    (display["Ecart_Mvt_Credit"].abs()    > 0.001) |
                    (display["Ecart_Solde_Debit"].abs()   > 0.001) |
                    (display["Ecart_Solde_Credit"].abs()  > 0.001)
                ]
                st.caption(f"{len(display)} compte(s) avec écart")
            st.dataframe(display.style.format(fmt).apply(_highlight_ecarts, axis=None),
                         use_container_width=True)

        with tab2:
            df_commun_disp = comp[comp["Compte"].isin(codes_communs)].copy()
            st.caption(f"{len(df_commun_disp)} comptes présents dans les deux fichiers")
            st.dataframe(df_commun_disp.style.format(fmt).apply(_highlight_ecarts, axis=None),
                         use_container_width=True)

        with tab3:
            if missing.empty:
                st.success("✅ Tous les comptes sont présents dans les deux fichiers.")
            else:
                only_a = missing[missing["Absent dans"] == LB]
                only_b = missing[missing["Absent dans"] == LA]
                st.info(
                    f"**{len(missing)} compte(s) manquant(s)** — "
                    f"🔴 {len(only_a)} absent(s) dans {LB} · "
                    f"🔵 {len(only_b)} absent(s) dans {LA}"
                )
                fmt_m = {c: "{:,.3f}" for c in missing.columns if missing[c].dtype == float}
                if not only_a.empty:
                    st.markdown(f"#### 🔴 Présents dans {LA} — Absents dans {LB}")
                    st.dataframe(only_a.drop(columns=["Présent dans", "Absent dans"])
                                 .reset_index(drop=True).style.format(fmt_m),
                                 use_container_width=True)
                if not only_b.empty:
                    st.markdown(f"#### 🔵 Présents dans {LB} — Absents dans {LA}")
                    st.dataframe(only_b.drop(columns=["Présent dans", "Absent dans"])
                                 .reset_index(drop=True).style.format(fmt_m),
                                 use_container_width=True)

        with tab4:
            st.dataframe(df1.style.format(
                {c: "{:,.3f}" for c in df1.columns if df1[c].dtype == float}
            ), use_container_width=True)

        with tab5:
            st.dataframe(df2.style.format(
                {c: "{:,.3f}" for c in df2.columns if df2[c].dtype == float}
            ), use_container_width=True)

        st.divider()
        st.download_button(
            label="📥 Télécharger Excel (5 onglets)",
            data=bg_build_excel(df1, df2, comp, missing, LA, LB),
            file_name="comparaison_balance_generale.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ══════════════════════════════════════════════════════════════════════════════
# MODULE 4 — GRAND LIVRE DÉTAILLÉ (transactions par compte, séparateur |)
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📗 Grand Livre Détaillé":

    st.title("📗 Comparaison Grand Livre Détaillé")

    @st.cache_data
    def parse_grand_livre_detail(file_bytes: bytes, label: str = "fichier") -> pd.DataFrame:
        """
        Format :
          compte XXXXXXXX DESCRIPTION
          DD/MM/YYYY|Référence|Type|Libelle|Débit|Crédit|Solde
        """
        lines = file_bytes.decode("utf-8", errors="ignore").splitlines()
        lines = [l.replace("\r", "") for l in lines]
        rows = []
        current_code = current_desc = ""

        for line in lines:
            m = re.match(r"^compte\s+(\d+)\s+(.*)", line, re.IGNORECASE)
            if m:
                current_code = m.group(1).strip()
                current_desc = m.group(2).strip()
                continue
            if re.match(r"^\d{2}/\d{2}/\d{4}\|", line):
                parts = line.split("|")
                if len(parts) < 6:
                    continue
                rows.append({
                    "Compte":      current_code,
                    "Description": current_desc,
                    "Date":        parts[0].strip(),
                    "Reference":   parts[1].strip(),
                    "Type":        parts[2].strip(),
                    "Libelle":     parts[3].strip(),
                    "Debit":       _to_float(parts[4]),
                    "Credit":      _to_float(parts[5]),
                    "Solde":       _to_float(parts[6]) if len(parts) > 6 else 0.0,
                })

        if not rows:
            st.warning(f"⚠️ **{label}** : aucune transaction détectée.")
            return pd.DataFrame(columns=["Compte", "Description", "Date", "Reference",
                                          "Type", "Libelle", "Debit", "Credit", "Solde"])
        df = pd.DataFrame(rows)
        for col in ["Debit", "Credit", "Solde"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
        return df

    def gld_compute_missing(df1, df2, la, lb):
        all_comptes = (pd.concat([df1[["Compte", "Description"]],
                                   df2[["Compte", "Description"]]])
                       .drop_duplicates("Compte").sort_values("Compte"))
        records = []
        for _, row in all_comptes.iterrows():
            code = row["Compte"]
            desc = row["Description"]
            refs1 = set(df1.loc[df1["Compte"] == code, "Reference"])
            refs2 = set(df2.loc[df2["Compte"] == code, "Reference"])
            for ref in sorted(refs1 - refs2):
                r = df1[(df1["Compte"] == code) & (df1["Reference"] == ref)].iloc[0]
                records.append({"Compte": code, "Description": desc,
                                "Reference": ref, "Présent dans": la, "Absent dans": lb,
                                "Date": r["Date"], "Type": r["Type"],
                                "Libelle": r["Libelle"],
                                "Debit": r["Debit"], "Credit": r["Credit"]})
            for ref in sorted(refs2 - refs1):
                r = df2[(df2["Compte"] == code) & (df2["Reference"] == ref)].iloc[0]
                records.append({"Compte": code, "Description": desc,
                                "Reference": ref, "Présent dans": lb, "Absent dans": la,
                                "Date": r["Date"], "Type": r["Type"],
                                "Libelle": r["Libelle"],
                                "Debit": r["Debit"], "Credit": r["Credit"]})
        if not records:
            return pd.DataFrame()
        return (pd.DataFrame(records)
                .sort_values(["Compte", "Présent dans", "Reference"])
                .reset_index(drop=True))

    def gld_build_excel(df1, df2, comp, missing, la, lb):
        wb      = Workbook()
        red_f   = PatternFill("solid", fgColor="FFCCCC")
        green_f = PatternFill("solid", fgColor="CCFFCC")

        ws1 = wb.active
        _style_sheet(ws1, comp, "Comparaison", "1F4E79")
        _excel_color_ecarts(ws1, comp, red_f, green_f)

        if not missing.empty:
            _excel_missing_sheet(wb, missing,
                                  ["Reference", "Présent dans", "Absent dans",
                                   "Date", "Type", "Libelle", "Debit", "Credit"],
                                  "Références manquantes", "C00000", la, lb,
                                  "Compte", "Description")

        ws3 = wb.create_sheet()
        _style_sheet(ws3, df1, f"Détail {la}", "2E75B6")
        ws4 = wb.create_sheet()
        _style_sheet(ws4, df2, f"Détail {lb}", "70AD47")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Upload ──────────────────────────────────────────────────────────────
    c1, c2 = st.columns(2)
    with c1:
        f1 = st.file_uploader("📂 Fichier A (Grand Livre)", type=["txt"], key="gld_f1")
        LA = st.text_input("Nom du fichier A", value="Fichier A", key="gld_la")
    with c2:
        f2 = st.file_uploader("📂 Fichier B (Grand Livre)", type=["txt"], key="gld_f2")
        LB = st.text_input("Nom du fichier B", value="Fichier B", key="gld_lb")

    if f1 and f2:
        with st.spinner("Parsing en cours..."):
            df1 = parse_grand_livre_detail(f1.read(), LA)
            df2 = parse_grand_livre_detail(f2.read(), LB)

        if df1.empty or df2.empty:
            st.error("Impossible de parser un ou plusieurs fichiers.")
            st.stop()

        sa, sb = f"_{LA}", f"_{LB}"

        # Agrégation par compte
        agg1 = (df1.groupby(["Compte", "Description"])[["Debit", "Credit"]]
                .sum().reset_index())
        solde1 = (df1.sort_values(["Compte", "Date"])
                  .groupby("Compte")["Solde"].last().reset_index()
                  .rename(columns={"Solde": "Solde_Final"}))
        agg1 = pd.merge(agg1, solde1, on="Compte", how="left")

        agg2 = (df2.groupby(["Compte", "Description"])[["Debit", "Credit"]]
                .sum().reset_index())
        solde2 = (df2.sort_values(["Compte", "Date"])
                  .groupby("Compte")["Solde"].last().reset_index()
                  .rename(columns={"Solde": "Solde_Final"}))
        agg2 = pd.merge(agg2, solde2, on="Compte", how="left")

        comp = pd.merge(agg1, agg2, on="Compte", how="outer", suffixes=(sa, sb))
        comp["Description"] = comp[f"Description{sa}"].fillna(comp[f"Description{sb}"])
        comp = comp.drop(columns=[f"Description{sa}", f"Description{sb}"]).fillna(0)
        comp = comp[["Compte", "Description",
                     f"Debit{sa}", f"Credit{sa}", f"Solde_Final{sa}",
                     f"Debit{sb}", f"Credit{sb}", f"Solde_Final{sb}"]]
        comp["Ecart_Debit"]       = comp[f"Debit{sb}"]       - comp[f"Debit{sa}"]
        comp["Ecart_Credit"]      = comp[f"Credit{sb}"]      - comp[f"Credit{sa}"]
        comp["Ecart_Solde_Final"] = comp[f"Solde_Final{sb}"] - comp[f"Solde_Final{sa}"]
        comp = comp.sort_values("Compte").reset_index(drop=True)

        missing = gld_compute_missing(df1, df2, LA, LB)

        nb_ecart = len(comp[
            (comp["Ecart_Debit"].abs()       > 0.001) |
            (comp["Ecart_Credit"].abs()      > 0.001) |
            (comp["Ecart_Solde_Final"].abs() > 0.001)
        ])

        k1, k2, k3, k4, k5, k6 = st.columns(6)
        k1.metric(f"Comptes {LA}", df1["Compte"].nunique())
        k2.metric(f"Comptes {LB}", df2["Compte"].nunique())
        k3.metric(f"Lignes {LA}", len(df1))
        k4.metric(f"Lignes {LB}", len(df2))
        k5.metric("Réf. manquantes", len(missing) if not missing.empty else 0,
                  delta="⚠️" if not missing.empty else None, delta_color="inverse")
        k6.metric("Comptes avec écart", nb_ecart,
                  delta="⚠️" if nb_ecart > 0 else None, delta_color="inverse")

        fmt = {c: "{:,.3f}" for c in comp.columns if comp[c].dtype == float}

        tab1, tab2, tab3, tab4 = st.tabs([
            "📊 Comparaison agrégée",
            "🔍 Références manquantes",
            f"📄 Détail {LA}",
            f"📄 Détail {LB}",
        ])

        with tab1:
            fc1, fc2 = st.columns([2, 1])
            with fc2:
                only_ecart = st.toggle("Afficher uniquement les écarts", value=False,
                                        key="gld_toggle")
            display = comp.copy()
            if only_ecart:
                display = display[
                    (display["Ecart_Debit"].abs()       > 0.001) |
                    (display["Ecart_Credit"].abs()      > 0.001) |
                    (display["Ecart_Solde_Final"].abs() > 0.001)
                ]
                st.caption(f"{len(display)} compte(s) avec écart")
            st.dataframe(display.style.format(fmt).apply(_highlight_ecarts, axis=None),
                         use_container_width=True)

        with tab2:
            if missing.empty:
                st.success("✅ Aucune référence manquante.")
            else:
                st.info(f"**{len(missing)} référence(s) manquante(s)** sur "
                        f"**{missing['Compte'].nunique()} compte(s)**")
                for compte_code, grp in missing.groupby("Compte", sort=True):
                    desc   = grp["Description"].iloc[0]
                    only_a = grp[grp["Absent dans"] == LB]
                    only_b = grp[grp["Absent dans"] == LA]
                    label  = (f"**{compte_code}** – {desc}  "
                              + (f"🔴 {len(only_a)} absent(s) dans {LB}  " if not only_a.empty else "")
                              + (f"🔵 {len(only_b)} absent(s) dans {LA}"   if not only_b.empty else ""))
                    with st.expander(label, expanded=False):
                        cols_d = ["Reference", "Date", "Type", "Libelle", "Debit", "Credit"]
                        fmt_m  = {"Debit": "{:,.3f}", "Credit": "{:,.3f}"}
                        if not only_a.empty:
                            st.markdown(f"🔴 **Présents dans {LA} — Absents dans {LB}**")
                            st.dataframe(only_a[cols_d].reset_index(drop=True).style.format(fmt_m),
                                         use_container_width=True)
                        if not only_b.empty:
                            st.markdown(f"🔵 **Présents dans {LB} — Absents dans {LA}**")
                            st.dataframe(only_b[cols_d].reset_index(drop=True).style.format(fmt_m),
                                         use_container_width=True)

        with tab3:
            comptes_a = ["Tous"] + sorted(df1["Compte"].unique().tolist())
            sel_a = st.selectbox(f"Filtrer par compte ({LA})", comptes_a, key="gld_sel_a")
            disp1 = df1 if sel_a == "Tous" else df1[df1["Compte"] == sel_a]
            st.caption(f"{len(disp1)} ligne(s)")
            st.dataframe(disp1.style.format(
                {c: "{:,.3f}" for c in disp1.columns if disp1[c].dtype == float}
            ), use_container_width=True)

        with tab4:
            comptes_b = ["Tous"] + sorted(df2["Compte"].unique().tolist())
            sel_b = st.selectbox(f"Filtrer par compte ({LB})", comptes_b, key="gld_sel_b")
            disp2 = df2 if sel_b == "Tous" else df2[df2["Compte"] == sel_b]
            st.caption(f"{len(disp2)} ligne(s)")
            st.dataframe(disp2.style.format(
                {c: "{:,.3f}" for c in disp2.columns if disp2[c].dtype == float}
            ), use_container_width=True)

        st.divider()
        st.download_button(
            label="📥 Télécharger Excel (4 onglets)",
            data=gld_build_excel(df1, df2, comp, missing, LA, LB),
            file_name="comparaison_grand_livre_detail.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
